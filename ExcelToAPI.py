from flask_restful import Api, Resource, reqparse
from flask import Flask, render_template, make_response
from pathlib import Path
import pandas as pd
import openpyxl
from werkzeug.datastructures import FileStorage
import os

# Intiation of the Flask app
app = Flask(__name__)
api = Api(app)

# Get the current path of this module
file_path = Path(__file__).parent.absolute()

# Setting the file path of the excel file that will be uploaded via the /upload api method
Excel_file_path = os.path.join(file_path, 'data.xlsx')


def search_value_in_col_idx(ws, search_string, col_idx=0):
    '''A function created to search in the sheet for a certain string at a certain column
    Arguments:
    ws = Sheet you want to search in
    Search_string = The string you want to search for
    col_idx = the column you want to search in (default=0[first column] )'''

    for row in range(1, ws.max_row + 1):
        if str(ws[row][col_idx].value) == search_string:
            return row
    return None


class WikiApi(Resource):
    '''The Class is responsible for all CRUD operations on the Excel sheet'''

    def get(self):
        '''This method reads all data from the excel sheet named data.xlsx and return it back in JSON'''

        data = pd.read_excel(Excel_file_path)
        data = data.to_dict('records')
        return {'data': data}, 200

    def post(self):
        '''This method creates a new entry in the data.xlsx and saves it'''

        parser = reqparse.RequestParser()
        parser.add_argument('novel', required=True)
        parser.add_argument('author', required=True)
        parser.add_argument('country', required=True)
        args = parser.parse_args()
        wb = openpyxl.load_workbook(filename=Excel_file_path)
        sheet = wb['Sheet1']
        row_count = sheet.max_row
        new_row = [row_count, args['novel'], args['author'], args['country']]
        sheet.append(new_row)
        wb.save(Excel_file_path)
        return {'message': 'Record added successfully.'}, 201

    def put(self):
        '''This method edits an existing entry in the data.xlsx and saves it/create it if not found'''

        parser = reqparse.RequestParser()
        parser.add_argument('bookno', required=True, type=int)
        parser.add_argument('novel', required=True)
        parser.add_argument('author', required=True)
        parser.add_argument('country', required=True)
        args = parser.parse_args()
        wb = openpyxl.load_workbook(filename=Excel_file_path)
        sheet = wb['Sheet1']
        idx = search_value_in_col_idx(sheet, str(args['bookno']))
        if idx:
            sheet.delete_rows(idx=idx)
            sheet.insert_rows(idx=idx)
            sheet.cell(column=1, row=idx, value=str(args['bookno']))
            sheet.cell(column=2, row=idx, value=str(args['novel']))
            sheet.cell(column=3, row=idx, value=str(args['author']))
            sheet.cell(column=4, row=idx, value=str(args['country']))
            wb.save(Excel_file_path)
            return {'message': 'Record Update successfully.'}, 200
        else:
            new_row = [str(args['bookno']), args['novel'],
                       args['author'], args['country']]
            sheet.append(new_row)
            wb.save(Excel_file_path)
            return {'message': 'No Match Found, Record Created Successfully'}, 201

    def delete(self):
        '''This method deletes an existing entry in the data.xlsx'''

        parser = reqparse.RequestParser()
        parser.add_argument('bookno', required=True, type=int)
        args = parser.parse_args()
        wb = openpyxl.load_workbook(filename=Excel_file_path)
        sheet = wb['Sheet1']
        idx = search_value_in_col_idx(sheet, str(args['bookno']))
        if idx:
            sheet.delete_rows(idx=idx)
            wb.save(Excel_file_path)
            return {'message': 'Record deleted successfully.'}, 200
        else:
            wb.close()
            return {'message': 'No record found matching this ID'}, 404


class UploadExcel(Resource):
    '''The Class is responsible for uploading the xlsx file generated from the wiki parser'''

    def post(self):
        '''The method uploads the excel file, renames it as data.xlsx and save it'''

        parse = reqparse.RequestParser()
        parse.add_argument('file', type=FileStorage, location='files')
        args = parse.parse_args()
        image_file = args['file']
        image_file.save(Excel_file_path)
        return {'message': 'File Uploaded Successfully, You can now start accessing the API via /api'}, 200


class HowTo(Resource):
    '''The Class is responsible showing the API documentation'''

    def get(self):
        '''The method renders an index.html template and returns it back in an HTTP response'''

        headers = {'Content-Type': 'text/html'}
        return make_response(render_template('index.html'), 200, headers)


# Add URL endpoints
api.add_resource(WikiApi, '/api')
api.add_resource(UploadExcel, '/upload')
api.add_resource(HowTo, '/')


if __name__ == '__main__':
    app.run()
